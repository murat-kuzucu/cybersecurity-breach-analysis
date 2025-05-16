import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

# Stil ve biçimlendirmeler için sabitler
TITLE_FONT = Font(name='Calibri', size=14, bold=True, color="00004C99")
HEADER_FONT = Font(name='Calibri', size=12, bold=True, color="000000")
SUBHEADER_FONT = Font(name='Calibri', size=11, bold=True, color="00333333")
QUESTION_FONT = Font(name='Calibri', size=10, bold=True, color="00000000")
OPTION_FONT = Font(name='Calibri', size=10, color="00000000")
HEADER_FILL = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")
SECTION_FILL = PatternFill(start_color="00E6E6E6", end_color="00E6E6E6", fill_type="solid")
BORDER_STYLE = Border(
    left=Side(style='thin', color="00000000"),
    right=Side(style='thin', color="00000000"),
    top=Side(style='thin', color="00000000"),
    bottom=Side(style='thin', color="00000000")
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
WRAP_TEXT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

def add_questionnaire_to_excel():
    """Anket sorularını Excel raporuna ekler"""
    
    # Excel dosyasını aç
    excelPath = 'Cybersecurity_Analysis_Report.xlsx'
    
    try:
        workbook = openpyxl.load_workbook(excelPath)
    except FileNotFoundError:
        print(f"{excelPath} bulunamadı.")
        return
    
    # Yeni çalışma sayfası oluştur
    if 'Questionnaire' in workbook.sheetnames:
        sheet = workbook['Questionnaire']
        # Sayfayı temizle
        for row in sheet.rows:
            for cell in row:
                cell.value = None
    else:
        sheet = workbook.create_sheet('Questionnaire')
    
    # Başlık ekle
    sheet['A1'] = 'MODERN CYBERSECURITY BREACH ANALYSIS - QUESTIONNAIRE'
    sheet['A1'].font = TITLE_FONT
    sheet.merge_cells('A1:G1')
    sheet['A1'].alignment = CENTER_ALIGN
    
    # Anket bölümlerini ve sorularını manuel olarak ekle
    currentRow = 3
    
    # Ana Araştırma Sorusu bölümü
    sheet[f'A{currentRow}'] = 'PRIMARY RESEARCH QUESTION'
    sheet[f'A{currentRow}'].font = HEADER_FONT
    sheet[f'A{currentRow}'].fill = HEADER_FILL
    sheet.merge_cells(f'A{currentRow}:G{currentRow}')
    currentRow += 1
    
    sheet[f'A{currentRow}'] = '"What are the most significant factors affecting financial losses and resolution times in cybersecurity breaches across different industries and regions in the current threat landscape?"'
    sheet[f'A{currentRow}'].alignment = WRAP_TEXT_ALIGN
    sheet.merge_cells(f'A{currentRow}:G{currentRow}')
    currentRow += 2
    
    # İkincil Araştırma Soruları
    sheet[f'A{currentRow}'] = 'SECONDARY RESEARCH QUESTIONS'
    sheet[f'A{currentRow}'].font = HEADER_FONT
    sheet[f'A{currentRow}'].fill = HEADER_FILL
    sheet.merge_cells(f'A{currentRow}:G{currentRow}')
    currentRow += 1
    
    secondaryQuestions = [
        '"How have cloud-based security solutions impacted the effectiveness of breach prevention and response?"',
        '"What role does Zero Trust Architecture play in mitigating the financial impact of breaches?"',
        '"How do remote/hybrid work environments affect an organization\'s cybersecurity posture and breach response capabilities?"'
    ]
    
    for question in secondaryQuestions:
        sheet[f'A{currentRow}'] = question
        sheet[f'A{currentRow}'].alignment = WRAP_TEXT_ALIGN
        sheet.merge_cells(f'A{currentRow}:G{currentRow}')
        currentRow += 1
    
    currentRow += 1
    
    # Hedef Kitle
    sheet[f'A{currentRow}'] = 'TARGET AUDIENCE'
    sheet[f'A{currentRow}'].font = HEADER_FONT
    sheet[f'A{currentRow}'].fill = HEADER_FILL
    sheet.merge_cells(f'A{currentRow}:G{currentRow}')
    currentRow += 1
    
    audiences = [
        'CISOs and security executives',
        'IT security professionals and security architects',
        'Organization security managers',
        'Cybersecurity analysts and threat hunters',
        'IT decision makers and risk managers',
        'DevSecOps practitioners',
        'Security operations center (SOC) team members'
    ]
    
    for audience in audiences:
        sheet[f'B{currentRow}'] = '☐'
        sheet[f'B{currentRow}'].alignment = CENTER_ALIGN
        
        sheet[f'C{currentRow}'] = audience
        sheet[f'C{currentRow}'].alignment = LEFT_ALIGN
        sheet.merge_cells(f'C{currentRow}:G{currentRow}')
        currentRow += 1
    
    currentRow += 1
    
    # Bölüm 1: Kuruluş Profili
    sheet[f'A{currentRow}'] = 'SECTION 1: ORGANIZATION PROFILE'
    sheet[f'A{currentRow}'].font = SUBHEADER_FONT
    sheet[f'A{currentRow}'].fill = SECTION_FILL
    sheet.merge_cells(f'A{currentRow}:G{currentRow}')
    currentRow += 1
    
    # Soru 1
    sheet[f'A{currentRow}'] = '1'
    sheet[f'A{currentRow}'].font = QUESTION_FONT
    sheet[f'A{currentRow}'].alignment = CENTER_ALIGN
    
    sheet[f'B{currentRow}'] = 'What industry does your organization primarily operate in?'
    sheet[f'B{currentRow}'].font = QUESTION_FONT
    sheet[f'B{currentRow}'].alignment = LEFT_ALIGN
    sheet.merge_cells(f'B{currentRow}:G{currentRow}')
    currentRow += 1
    
    industries = [
        'Banking/Finance/FinTech',
        'Healthcare/Medical Devices',
        'Education/EdTech',
        'Retail/E-commerce',
        'IT/Technology/SaaS',
        'Government/Public Sector',
        'Energy/Utilities/Critical Infrastructure',
        'Manufacturing/Industrial IoT',
        'Telecommunications/ISP',
        'Other (please specify): _______________'
    ]
    
    for industry in industries:
        sheet[f'B{currentRow}'] = '☐'
        sheet[f'B{currentRow}'].alignment = CENTER_ALIGN
        
        sheet[f'C{currentRow}'] = industry
        sheet[f'C{currentRow}'].alignment = LEFT_ALIGN
        sheet.merge_cells(f'C{currentRow}:G{currentRow}')
        currentRow += 1
    
    currentRow += 1
    
    # Soru 2
    sheet[f'A{currentRow}'] = '2'
    sheet[f'A{currentRow}'].font = QUESTION_FONT
    sheet[f'A{currentRow}'].alignment = CENTER_ALIGN
    
    sheet[f'B{currentRow}'] = 'What is the approximate size of your organization?'
    sheet[f'B{currentRow}'].font = QUESTION_FONT
    sheet[f'B{currentRow}'].alignment = LEFT_ALIGN
    sheet.merge_cells(f'B{currentRow}:G{currentRow}')
    currentRow += 1
    
    sizes = [
        'Small (1-50 employees)',
        'Medium (51-500 employees)',
        'Large (501-5000 employees)',
        'Enterprise (5000+ employees)'
    ]
    
    for size in sizes:
        sheet[f'B{currentRow}'] = '☐'
        sheet[f'B{currentRow}'].alignment = CENTER_ALIGN
        
        sheet[f'C{currentRow}'] = size
        sheet[f'C{currentRow}'].alignment = LEFT_ALIGN
        sheet.merge_cells(f'C{currentRow}:G{currentRow}')
        currentRow += 1
    
    currentRow += 1
    
    # Bölüm 2: Güvenlik İhlali Deneyimi
    sheet[f'A{currentRow}'] = 'SECTION 2: SECURITY BREACH EXPERIENCE'
    sheet[f'A{currentRow}'].font = SUBHEADER_FONT
    sheet[f'A{currentRow}'].fill = SECTION_FILL
    sheet.merge_cells(f'A{currentRow}:G{currentRow}')
    currentRow += 1
    
    # Soru 6
    sheet[f'A{currentRow}'] = '6'
    sheet[f'A{currentRow}'].font = QUESTION_FONT
    sheet[f'A{currentRow}'].alignment = CENTER_ALIGN
    
    sheet[f'B{currentRow}'] = 'Has your organization experienced a significant cybersecurity breach in the past 3 years?'
    sheet[f'B{currentRow}'].font = QUESTION_FONT
    sheet[f'B{currentRow}'].alignment = LEFT_ALIGN
    sheet.merge_cells(f'B{currentRow}:G{currentRow}')
    currentRow += 1
    
    options = [
        'Yes, multiple breaches',
        'Yes, one breach',
        'No',
        'Unsure',
        'Prefer not to disclose'
    ]
    
    for option in options:
        sheet[f'B{currentRow}'] = '☐'
        sheet[f'B{currentRow}'].alignment = CENTER_ALIGN
        
        sheet[f'C{currentRow}'] = option
        sheet[f'C{currentRow}'].alignment = LEFT_ALIGN
        sheet.merge_cells(f'C{currentRow}:G{currentRow}')
        currentRow += 1
    
    currentRow += 1
    
    # Soru 7
    sheet[f'A{currentRow}'] = '7'
    sheet[f'A{currentRow}'].font = QUESTION_FONT
    sheet[f'A{currentRow}'].alignment = CENTER_ALIGN
    
    sheet[f'B{currentRow}'] = 'If yes, what type(s) of attack did your organization experience? (Select all that apply)'
    sheet[f'B{currentRow}'].font = QUESTION_FONT
    sheet[f'B{currentRow}'].alignment = LEFT_ALIGN
    sheet.merge_cells(f'B{currentRow}:G{currentRow}')
    currentRow += 1
    
    attackTypes = [
        'Phishing/Spear-phishing/Business Email Compromise',
        'Ransomware/Double-extortion ransomware',
        'DDoS',
        'Supply chain attack',
        'Cloud configuration exploitation',
        'Zero-day exploit',
        'Social Engineering',
        'Man-in-the-Middle',
        'SQL Injection/Web application attacks',
        'Insider threat',
        'Other (please specify): _______________'
    ]
    
    for attackType in attackTypes:
        sheet[f'B{currentRow}'] = '☐'
        sheet[f'B{currentRow}'].alignment = CENTER_ALIGN
        
        sheet[f'C{currentRow}'] = attackType
        sheet[f'C{currentRow}'].alignment = LEFT_ALIGN
        sheet.merge_cells(f'C{currentRow}:G{currentRow}')
        currentRow += 1
    
    currentRow += 1
    
    # Finansal Etki Sorusu
    sheet[f'A{currentRow}'] = '9'
    sheet[f'A{currentRow}'].font = QUESTION_FONT
    sheet[f'A{currentRow}'].alignment = CENTER_ALIGN
    
    sheet[f'B{currentRow}'] = 'What was the approximate total financial impact from the most significant breach? (in USD)'
    sheet[f'B{currentRow}'].font = QUESTION_FONT
    sheet[f'B{currentRow}'].alignment = LEFT_ALIGN
    sheet.merge_cells(f'B{currentRow}:G{currentRow}')
    currentRow += 1
    
    financialImpacts = [
        'Less than $10,000',
        '$10,000 - $50,000',
        '$50,001 - $100,000',
        '$100,001 - $500,000',
        '$500,001 - $1 million',
        '$1 million - $5 million',
        '$5 million - $10 million',
        'More than $10 million',
        'Unknown/Still calculating',
        'Prefer not to disclose'
    ]
    
    for impact in financialImpacts:
        sheet[f'B{currentRow}'] = '☐'
        sheet[f'B{currentRow}'].alignment = CENTER_ALIGN
        
        sheet[f'C{currentRow}'] = impact
        sheet[f'C{currentRow}'].alignment = LEFT_ALIGN
        sheet.merge_cells(f'C{currentRow}:G{currentRow}')
        currentRow += 1
    
    # Sütun genişliklerini ayarla
    sheet.column_dimensions['A'].width = 8
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 40
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['G'].width = 20
    
    # Sayfa formatını ayarla
    sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToPage = True
    sheet.page_setup.fitToHeight = 0
    sheet.page_setup.fitToWidth = 1
    
    # Not ekle
    noteRow = currentRow + 2
    sheet[f'A{noteRow}'] = 'Note: This is a simplified excerpt of the full questionnaire. The complete version contains additional questions on security vulnerabilities, resolution times, and post-breach responses.'
    sheet[f'A{noteRow}'].font = Font(italic=True)
    sheet.merge_cells(f'A{noteRow}:G{noteRow}')
    sheet[f'A{noteRow}'].alignment = WRAP_TEXT_ALIGN
    
    # Excel dosyasını kaydet
    workbook.save('Cybersecurity_Analysis_Report_with_Questionnaire.xlsx')
    print(f"Anket soruları 'Cybersecurity_Analysis_Report_with_Questionnaire.xlsx' dosyasına eklendi.")

if __name__ == "__main__":
    add_questionnaire_to_excel()
