import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
from pathlib import Path
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, Rule
import io

def create_excel_reports():
    """Create comprehensive Excel reports from cybersecurity breach data"""
    
    # Veri yolu ayarları
    filePath = Path('dataset-onExcel.xlsx')
    outputPath = Path('Cybersecurity_Analysis_Report.xlsx')
    outputDir = Path('analysis_results')
    outputDir.mkdir(exist_ok=True)
    
    # Veriyi yükle
    dataFrame = pd.read_excel(filePath)
    
    # Excel dosyası oluşturma
    writer = pd.ExcelWriter('Cybersecurity_Analysis_Report.xlsx', engine='openpyxl')
    
    # Ana veri setini ilk sayfaya yaz
    dataFrame.to_excel(writer, sheet_name='Raw Data', index=False)
    
    # Özet istatistikler sayfası
    create_summary_sheet(dataFrame, writer)
    
    # Finansal etki analizi sayfası
    create_financial_analysis_sheet(dataFrame, writer)
    
    # Saldırı türleri analiz sayfası
    create_attack_analysis_sheet(dataFrame, writer)
    
    # Çözüm süreleri analiz sayfası
    create_resolution_analysis_sheet(dataFrame, writer)
    
    # Güvenlik açıkları analiz sayfası
    create_vulnerability_analysis_sheet(dataFrame, writer)
    
    # Saldırı kaynakları analiz sayfası
    create_attack_source_sheet(dataFrame, writer)
    
    # Dashboard sayfası
    create_dashboard_sheet(dataFrame, writer)
    
    # Excel dosyasını kaydet
    writer.close()
    
    print(f"Excel raporu oluşturuldu: {outputPath}")
    return outputPath

def create_summary_sheet(df, writer):
    """Özet istatistikler sayfası oluştur"""
    
    # Özet DataFrame'ler oluştur
    summaryStats = df.describe().T
    missingValues = pd.DataFrame(df.isnull().sum(), columns=['Missing Values'])
    attackCounts = df['Attack Type'].value_counts().reset_index()
    attackCounts.columns = ['Attack Type', 'Count']
    yearCounts = df['Year'].value_counts().sort_index().reset_index()
    yearCounts.columns = ['Year', 'Count']
    
    # Veriyi Excel'e yaz
    summaryStats.to_excel(writer, sheet_name='Summary', startrow=1, startcol=0)
    missingValues.to_excel(writer, sheet_name='Summary', startrow=len(summaryStats) + 5, startcol=0)
    attackCounts.to_excel(writer, sheet_name='Summary', startrow=1, startcol=15, index=False)
    yearCounts.to_excel(writer, sheet_name='Summary', startrow=len(attackCounts) + 5, startcol=15, index=False)
    
    # Çalışma sayfasını al
    worksheet = writer.sheets['Summary']
    
    # Başlıklar
    worksheet['A1'] = 'DESCRIPTIVE STATISTICS'
    worksheet['P1'] = 'ATTACK TYPE DISTRIBUTION'
    worksheet['P{}'.format(len(attackCounts) + 5)] = 'YEARLY ATTACK DISTRIBUTION'
    worksheet['A{}'.format(len(summaryStats) + 5)] = 'MISSING VALUES'
    
    # Başlıkları biçimlendir
    for cell in [worksheet['A1'], worksheet['P1'], worksheet['P{}'.format(len(attackCounts) + 5)], worksheet['A{}'.format(len(summaryStats) + 5)]]:
        cell.font = Font(bold=True, size=14)
    
    # Saldırı türü için pasta grafiği oluştur
    pieChart = openpyxl.chart.PieChart()
    labels = Reference(worksheet, min_col=16, min_row=2, max_row=len(attackCounts) + 1)
    data = Reference(worksheet, min_col=17, min_row=1, max_row=len(attackCounts) + 1)
    pieChart.add_data(data, titles_from_data=True)
    pieChart.set_categories(labels)
    pieChart.title = "Attack Type Distribution"
    pieChart.height = 15
    pieChart.width = 20
    
    # Pasta grafiğini sayfaya ekle
    worksheet.add_chart(pieChart, "O{}".format(len(attackCounts) + 15))
    
    # Yıllık saldırı grafiği
    barChart = openpyxl.chart.BarChart()
    labels = Reference(worksheet, min_col=16, min_row=len(attackCounts) + 6, max_row=len(attackCounts) + 6 + len(yearCounts) - 1)
    data = Reference(worksheet, min_col=17, min_row=len(attackCounts) + 5, max_row=len(attackCounts) + 5 + len(yearCounts))
    barChart.add_data(data, titles_from_data=True)
    barChart.set_categories(labels)
    barChart.title = "Yearly Attack Distribution"
    barChart.height = 15
    barChart.width = 20
    
    # Çubuk grafiğini sayfaya ekle
    worksheet.add_chart(barChart, "O{}".format(len(attackCounts) + len(yearCounts) + 25))
    
    # Sütun genişliklerini ayarla
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width

def create_financial_analysis_sheet(df, writer):
    """Finansal etki analizi sayfası oluştur"""
    
    # Sektör bazında ortalama finansal kayıp
    industryLoss = df.groupby('Target Industry')['Financial Loss (in Million $)'].agg(['mean', 'sum', 'count']).sort_values('sum', ascending=False).reset_index()
    
    # Yıllara göre finansal kayıp trendi
    yearlyLoss = df.groupby('Year')['Financial Loss (in Million $)'].mean().reset_index()
    
    # Sektörlere göre yıllık trend (top 5)
    topIndustries = industryLoss.nlargest(5, 'sum')['Target Industry'].tolist()
    timeTrend = pd.pivot_table(
        df[df['Target Industry'].isin(topIndustries)],
        values='Financial Loss (in Million $)',
        index='Year',
        columns='Target Industry',
        aggfunc='mean'
    ).reset_index()
    
    # Excel'e yaz
    industryLoss.to_excel(writer, sheet_name='Financial Analysis', startrow=1, startcol=0, index=False)
    yearlyLoss.to_excel(writer, sheet_name='Financial Analysis', startrow=1, startcol=6, index=False)
    timeTrend.to_excel(writer, sheet_name='Financial Analysis', startrow=1, startcol=10, index=False)
    
    # Çalışma sayfasını al
    worksheet = writer.sheets['Financial Analysis']
    
    # Başlıklar
    worksheet['A1'] = 'FINANCIAL LOSS BY INDUSTRY'
    worksheet['G1'] = 'FINANCIAL LOSS TREND OVER TIME'
    worksheet['K1'] = 'FINANCIAL LOSS BY TOP INDUSTRIES OVER TIME'
    
    # Başlıkları biçimlendir
    for cell in [worksheet['A1'], worksheet['G1'], worksheet['K1']]:
        cell.font = Font(bold=True, size=14)
    
    # Çubuk grafiği oluştur (Sektöre göre kayıp)
    barChart = openpyxl.chart.BarChart()
    labels = Reference(worksheet, min_col=1, min_row=2, max_row=len(industryLoss) + 1)
    data = Reference(worksheet, min_col=2, min_row=1, max_row=len(industryLoss) + 1)
    barChart.add_data(data, titles_from_data=True)
    barChart.set_categories(labels)
    barChart.title = "Average Financial Loss by Industry"
    barChart.y_axis.title = "Industry"
    barChart.x_axis.title = "Average Loss (Million $)"
    barChart.height = 15
    barChart.width = 20
    
    # Çubuk grafiğini sayfaya ekle
    worksheet.add_chart(barChart, "A{}".format(len(industryLoss) + 5))
    
    # Çizgi grafiği oluştur (Yıllara göre trend)
    lineChart = openpyxl.chart.LineChart()
    labels = Reference(worksheet, min_col=7, min_row=2, max_row=len(yearlyLoss) + 1)
    data = Reference(worksheet, min_col=8, min_row=1, max_row=len(yearlyLoss) + 1)
    lineChart.add_data(data, titles_from_data=True)
    lineChart.set_categories(labels)
    lineChart.title = "Average Financial Loss Over Time"
    lineChart.y_axis.title = "Average Loss (Million $)"
    lineChart.x_axis.title = "Year"
    lineChart.height = 15
    lineChart.width = 20
    
    # Çizgi grafiğini sayfaya ekle
    worksheet.add_chart(lineChart, "G{}".format(len(yearlyLoss) + 5))
    
    # Koşullu biçimlendirme (yüksek değerler için)
    redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    yellowFill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
    greenFill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
    
    # Veri aralığı
    dataRange = "B2:B{}".format(len(industryLoss) + 1)
    worksheet.conditional_formatting.add(
        dataRange,
        ColorScaleRule(start_type='percentile', start_value=0, start_color='FF63BE7B',
                      mid_type='percentile', mid_value=50, mid_color='FFFFEB84',
                      end_type='percentile', end_value=100, end_color='FFF8696B')
    )

def create_attack_analysis_sheet(df, writer):
    """Saldırı türleri analiz sayfası oluştur"""
    
    # Saldırı türleri dağılımı
    attackCounts = df['Attack Type'].value_counts().reset_index()
    attackCounts.columns = ['Attack Type', 'Count']
    
    # Yıllara göre saldırı türleri
    attackByYear = pd.pivot_table(
        df, 
        values='Financial Loss (in Million $)',
        index='Year',
        columns='Attack Type',
        aggfunc='mean'
    ).reset_index()
    
    # Ülkelere göre saldırı türleri
    attackByCountry = pd.pivot_table(
        df,
        values='Financial Loss (in Million $)',
        index='Country',
        columns='Attack Type',
        aggfunc='mean'
    ).reset_index()
    
    # Excel'e yaz
    attackCounts.to_excel(writer, sheet_name='Attack Analysis', startrow=1, startcol=0, index=False)
    attackByYear.to_excel(writer, sheet_name='Attack Analysis', startrow=1, startcol=4, index=False)
    attackByCountry.to_excel(writer, sheet_name='Attack Analysis', startrow=len(attackByYear) + 5, startcol=4, index=False)
    
    # Çalışma sayfasını al
    worksheet = writer.sheets['Attack Analysis']
    
    # Başlıklar
    worksheet['A1'] = 'ATTACK TYPE DISTRIBUTION'
    worksheet['E1'] = 'ATTACK TYPES BY YEAR (Avg. Financial Loss)'
    worksheet['E{}'.format(len(attackByYear) + 5)] = 'ATTACK TYPES BY COUNTRY (Avg. Financial Loss)'
    
    # Başlıkları biçimlendir
    for cell in [worksheet['A1'], worksheet['E1'], worksheet['E{}'.format(len(attackByYear) + 5)]]:
        cell.font = Font(bold=True, size=14)
    
    # Saldırı türleri için pasta grafiği oluştur
    pieChart = openpyxl.chart.PieChart()
    labels = Reference(worksheet, min_col=1, min_row=2, max_row=len(attackCounts) + 1)
    data = Reference(worksheet, min_col=2, min_row=1, max_row=len(attackCounts) + 1)
    pieChart.add_data(data, titles_from_data=True)
    pieChart.set_categories(labels)
    pieChart.title = "Attack Type Distribution"
    pieChart.height = 15
    pieChart.width = 20
    
    # Pasta grafiğini sayfaya ekle
    worksheet.add_chart(pieChart, "A{}".format(len(attackCounts) + 5))

def create_resolution_analysis_sheet(df, writer):
    """Çözüm süreleri analiz sayfası oluştur"""
    
    # Saldırı türüne göre ortalama çözüm süresi
    resolutionByAttack = df.groupby('Attack Type')['Incident Resolution Time (in Hours)'].agg(['mean', 'median', 'count']).sort_values('mean', ascending=False).reset_index()
    
    # Ülkeye göre çözüm süresi
    resolutionByCountry = df.groupby('Country')['Incident Resolution Time (in Hours)'].mean().sort_values(ascending=False).reset_index()
    
    # Çözüm süresi ile mali kayıp arasındaki ilişki
    pivotData = pd.pivot_table(
        df,
        values='Financial Loss (in Million $)',
        index=pd.cut(df['Incident Resolution Time (in Hours)'], bins=[0, 24, 48, 72, 96, float('inf')]),
        columns='Attack Type',
        aggfunc='mean'
    ).reset_index()
    pivotData.columns.name = None
    pivotData.rename(columns={pivotData.columns[0]: 'Resolution Time Range'}, inplace=True)
    
    # Excel'e yaz
    resolutionByAttack.to_excel(writer, sheet_name='Resolution Analysis', startrow=1, startcol=0, index=False)
    resolutionByCountry.to_excel(writer, sheet_name='Resolution Analysis', startrow=1, startcol=5, index=False)
    pivotData.to_excel(writer, sheet_name='Resolution Analysis', startrow=len(resolutionByAttack) + 5, startcol=0, index=False)
    
    # Çalışma sayfasını al
    worksheet = writer.sheets['Resolution Analysis']
    
    # Başlıklar
    worksheet['A1'] = 'RESOLUTION TIME BY ATTACK TYPE'
    worksheet['F1'] = 'RESOLUTION TIME BY COUNTRY'
    worksheet['A{}'.format(len(resolutionByAttack) + 5)] = 'FINANCIAL LOSS BY RESOLUTION TIME RANGE'
    
    # Başlıkları biçimlendir
    for cell in [worksheet['A1'], worksheet['F1'], worksheet['A{}'.format(len(resolutionByAttack) + 5)]]:
        cell.font = Font(bold=True, size=14)
    
    # Çubuk grafiği oluştur (Saldırı türüne göre çözüm süresi)
    barChart = openpyxl.chart.BarChart()
    labels = Reference(worksheet, min_col=1, min_row=2, max_row=len(resolutionByAttack) + 1)
    data = Reference(worksheet, min_col=2, min_row=1, max_row=len(resolutionByAttack) + 1)
    barChart.add_data(data, titles_from_data=True)
    barChart.set_categories(labels)
    barChart.title = "Average Resolution Time by Attack Type"
    barChart.y_axis.title = "Attack Type"
    barChart.x_axis.title = "Hours"
    barChart.height = 15
    barChart.width = 20
    
    # Çubuk grafiğini sayfaya ekle
    worksheet.add_chart(barChart, "A{}".format(len(resolutionByAttack) + 15))

def create_vulnerability_analysis_sheet(df, writer):
    """Güvenlik açıkları analiz sayfası oluştur"""
    
    # Güvenlik açığı türüne göre finansal kayıp
    vulnLoss = df.groupby('Security Vulnerability Type')['Financial Loss (in Million $)'].agg(['mean', 'sum', 'count']).sort_values('sum', ascending=False).reset_index()
    
    # Savunma mekanizmasına göre finansal kayıp ve çözüm süresi
    defenseLoss = df.groupby('Defense Mechanism Used')['Financial Loss (in Million $)'].mean().sort_values().reset_index()
    defenseTime = df.groupby('Defense Mechanism Used')['Incident Resolution Time (in Hours)'].mean().sort_values().reset_index()
    
    # Excel'e yaz
    vulnLoss.to_excel(writer, sheet_name='Vulnerability Analysis', startrow=1, startcol=0, index=False)
    defenseLoss.to_excel(writer, sheet_name='Vulnerability Analysis', startrow=1, startcol=5, index=False)
    defenseTime.to_excel(writer, sheet_name='Vulnerability Analysis', startrow=1, startcol=8, index=False)
    
    # Çalışma sayfasını al
    worksheet = writer.sheets['Vulnerability Analysis']
    
    # Başlıklar
    worksheet['A1'] = 'FINANCIAL LOSS BY VULNERABILITY TYPE'
    worksheet['F1'] = 'FINANCIAL LOSS BY DEFENSE MECHANISM'
    worksheet['I1'] = 'RESOLUTION TIME BY DEFENSE MECHANISM'
    
    # Başlıkları biçimlendir
    for cell in [worksheet['A1'], worksheet['F1'], worksheet['I1']]:
        cell.font = Font(bold=True, size=14)
    
    # Çubuk grafiği oluştur (Güvenlik açığına göre kayıp)
    barChart1 = openpyxl.chart.BarChart()
    labels = Reference(worksheet, min_col=1, min_row=2, max_row=len(vulnLoss) + 1)
    data = Reference(worksheet, min_col=2, min_row=1, max_row=len(vulnLoss) + 1)
    barChart1.add_data(data, titles_from_data=True)
    barChart1.set_categories(labels)
    barChart1.title = "Average Financial Loss by Vulnerability Type"
    barChart1.y_axis.title = "Vulnerability Type"
    barChart1.x_axis.title = "Average Loss (Million $)"
    barChart1.height = 15
    barChart1.width = 20
    
    # Çubuk grafiğini sayfaya ekle
    worksheet.add_chart(barChart1, "A{}".format(len(vulnLoss) + 5))
    
    # Çubuk grafiği oluştur (Savunma mekanizmasına göre kayıp)
    barChart2 = openpyxl.chart.BarChart()
    labels = Reference(worksheet, min_col=6, min_row=2, max_row=len(defenseLoss) + 1)
    data = Reference(worksheet, min_col=7, min_row=1, max_row=len(defenseLoss) + 1)
    barChart2.add_data(data, titles_from_data=True)
    barChart2.set_categories(labels)
    barChart2.title = "Average Financial Loss by Defense Mechanism"
    barChart2.y_axis.title = "Defense Mechanism"
    barChart2.x_axis.title = "Average Loss (Million $)"
    barChart2.height = 15
    barChart2.width = 20
    
    # Çubuk grafiğini sayfaya ekle
    worksheet.add_chart(barChart2, "F{}".format(len(defenseLoss) + 5))

def create_attack_source_sheet(df, writer):
    """Saldırı kaynakları analiz sayfası oluştur"""
    
    # Saldırı kaynağı dağılımı
    sourceCounts = df['Attack Source'].value_counts().reset_index()
    sourceCounts.columns = ['Attack Source', 'Count']
    
    # Saldırı kaynağı ve ülke
    sourceByCountry = pd.crosstab(df['Country'], df['Attack Source']).reset_index()
    
    # Saldırı kaynağı ve saldırı türü
    sourceByAttack = pd.crosstab(df['Attack Source'], df['Attack Type']).reset_index()
    
    # Excel'e yaz
    sourceCounts.to_excel(writer, sheet_name='Attack Sources', startrow=1, startcol=0, index=False)
    sourceByCountry.to_excel(writer, sheet_name='Attack Sources', startrow=1, startcol=4, index=False)
    sourceByAttack.to_excel(writer, sheet_name='Attack Sources', startrow=len(sourceByCountry) + 5, startcol=4, index=False)
    
    # Çalışma sayfasını al
    worksheet = writer.sheets['Attack Sources']
    
    # Başlıklar
    worksheet['A1'] = 'ATTACK SOURCE DISTRIBUTION'
    worksheet['E1'] = 'ATTACK SOURCES BY COUNTRY'
    worksheet['E{}'.format(len(sourceByCountry) + 5)] = 'ATTACK SOURCES BY ATTACK TYPE'
    
    # Başlıkları biçimlendir
    for cell in [worksheet['A1'], worksheet['E1'], worksheet['E{}'.format(len(sourceByCountry) + 5)]]:
        cell.font = Font(bold=True, size=14)
    
    # Pasta grafiği oluştur (Saldırı kaynağı dağılımı)
    pieChart = openpyxl.chart.PieChart()
    labels = Reference(worksheet, min_col=1, min_row=2, max_row=len(sourceCounts) + 1)
    data = Reference(worksheet, min_col=2, min_row=1, max_row=len(sourceCounts) + 1)
    pieChart.add_data(data, titles_from_data=True)
    pieChart.set_categories(labels)
    pieChart.title = "Attack Source Distribution"
    pieChart.height = 15
    pieChart.width = 20
    
    # Pasta grafiğini sayfaya ekle
    worksheet.add_chart(pieChart, "A{}".format(len(sourceCounts) + 5))

def create_dashboard_sheet(df, writer):
    """Özet dashboard sayfası oluştur"""
    
    # Çalışma sayfası oluştur
    if 'Dashboard' not in writer.sheets:
        writer.sheets['Dashboard'] = writer.book.create_sheet('Dashboard')
    worksheet = writer.sheets['Dashboard']
    
    # Başlık
    worksheet['A1'] = 'CYBERSECURITY BREACH ANALYSIS DASHBOARD'
    worksheet['A1'].font = Font(bold=True, size=18)
    
    # Üst kısım KPI'ları
    worksheet['A3'] = 'SUMMARY METRICS'
    worksheet['A3'].font = Font(bold=True, size=14)
    
    # Toplam finansal kayıp
    worksheet['A5'] = 'Total Financial Loss:'
    worksheet['B5'] = df['Financial Loss (in Million $)'].sum()
    worksheet['B5'].font = Font(bold=True, size=12, color="FF0000")
    
    # Etkilenen kullanıcı sayısı
    worksheet['A6'] = 'Total Affected Users:'
    worksheet['B6'] = df['Number of Affected Users'].sum()
    worksheet['B6'].font = Font(bold=True, size=12)
    
    # Ortalama çözüm süresi
    worksheet['A7'] = 'Average Resolution Time (Hours):'
    worksheet['B7'] = df['Incident Resolution Time (in Hours)'].mean()
    worksheet['B7'].font = Font(bold=True, size=12)
    
    # En yaygın saldırı türü
    worksheet['A8'] = 'Most Common Attack Type:'
    worksheet['B8'] = df['Attack Type'].value_counts().index[0]
    worksheet['B8'].font = Font(bold=True, size=12)
    
    # En çok etkilenen ülke
    worksheet['A9'] = 'Most Affected Country:'
    worksheet['B9'] = df.groupby('Country')['Financial Loss (in Million $)'].sum().sort_values(ascending=False).index[0]
    worksheet['B9'].font = Font(bold=True, size=12)
    
    # En çok etkilenen sektör
    worksheet['A10'] = 'Most Affected Industry:'
    worksheet['B10'] = df.groupby('Target Industry')['Financial Loss (in Million $)'].sum().sort_values(ascending=False).index[0]
    worksheet['B10'].font = Font(bold=True, size=12)
    
    # Ana sayfalara linkler
    worksheet['A12'] = 'NAVIGATE TO DETAILED ANALYSES:'
    worksheet['A12'].font = Font(bold=True, size=14)
    
    # Linkler
    links = [
        ('A14', 'Raw Data', 'Raw Data'),
        ('A15', 'Summary Statistics', 'Summary'),
        ('A16', 'Financial Analysis', 'Financial Analysis'),
        ('A17', 'Attack Analysis', 'Attack Analysis'),
        ('A18', 'Resolution Time Analysis', 'Resolution Analysis'),
        ('A19', 'Vulnerability Analysis', 'Vulnerability Analysis'),
        ('A20', 'Attack Sources', 'Attack Sources')
    ]
    
    for cell, text, sheet in links:
        worksheet[cell] = text
        worksheet[cell].font = Font(bold=True, color="0000FF", underline="single")
        # Bağlantı eklenebilir ancak openpyxl ile biraz karmaşık

    # Sütun genişliklerini ayarla
    worksheet.column_dimensions['A'].width = 30
    worksheet.column_dimensions['B'].width = 30

if __name__ == "__main__":
    excelPath = create_excel_reports()
    print(f"Excel raporu başarıyla oluşturuldu: {excelPath}")
